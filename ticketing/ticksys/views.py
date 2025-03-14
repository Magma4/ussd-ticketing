from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from .forms import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.utils import timezone
from django.db.models.functions import TruncMonth, ExtractMonth, ExtractDay, ExtractYear
from django.db.models import Count
from datetime import datetime
from django.template.loader import render_to_string
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from django.views import generic
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Sum
from bootstrap_modal_forms.generic import (
  BSModalCreateView,
  BSModalUpdateView,
  BSModalDeleteView
)
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from reportlab.lib import colors


# Create your views here.
class Index(generic.ListView):
    model = Ticket
    context_object_name = 'ticket'
    template_name = 'index.html'

# Create
class TicketCreateView(CreateView):
    model = Ticket
    form_class = TicketForm
    template_name = 'ticket_form_modal.html'
    success_url = reverse_lazy('index')
    success_message = 'Ticket created succesfully.'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request  # Pass request to the form
        return kwargs

    def form_valid(self, form):
        instance = form.save(commit=False)
        instance.user = self.request.user  # Assign the user to the instance
        instance.save()
        messages.success(self.request, "Ticket created")
        return super().form_valid(form)

class TicketUpdateView(UpdateView):
    model = Ticket
    form_class = TicketForm
    template_name = 'ticket_form_modal.html'  # Reuse the same template for consistency
    success_url = reverse_lazy('index')  # Redirect to the index page after successful update

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request  # Pass request to the form
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, "Ticket updated successfully.")
        return super().form_valid(form)

class TicketDeleteView(DeleteView):
    model = Ticket
    template_name = 'ticket_confirm_delete.html'  # Template for confirmation
    success_url = reverse_lazy('index')  # Redirect to the index page after successful deletion

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Ticket deleted successfully.")
        return super().delete(request, *args, **kwargs)
# Update
# class TicketUpdateView(BSModalUpdateView):
#     model = Ticket
#     template_name = 'dashboard/order_update.html'
#     form_class = OrderForm
#     success_message = 'Request was updated.'

#     def form_valid(self, form):
#         instance = form.save(commit=False)  # Save the form data without committing to the database yet
#         request_quantity = instance.request_quantity
#         stock_quantity = instance.item_name.get_total_quantity()  # Use the method to get the total quantity

#         if request_quantity <= stock_quantity:
#             messages.success(self.request, "Request updated successfully")
#             return super().form_valid(form)  # This saves the form and redirects to success_url
#         else:
#             messages.error(self.request, "Requested quantity exceeds available stock")
#             # Redirecting to the same page if the form is not valid
#             return HttpResponseRedirect(self.request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

#     def get_context_data(self, **kwargs):
#         context = super(OrderUpdateView, self).get_context_data(**kwargs)
#         context['orders'] = Order.objects.all()  # Adding orders to the context
#         return context

#     def get_success_url(self):
#         # Check if the user is a superuser
#         if self.request.user.is_superuser:
#             return reverse('view-request')  # Redirect superusers to the dashboard
#         else:
#             return reverse('dashboard')


# # Delete
# class OrderDeleteView(BSModalDeleteView):
#     model = Order
#     template_name = 'dashboard/order_delete.html'
#     success_message = 'Request was deleted.'


#     def get_success_url(self):
#         # Check if the user is a superuser
#         if self.request.user.is_superuser:
#             return reverse('view-request')  # Redirect superusers to the dashboard
#         else:
#             return reverse('dashboard')
# Create your views here.

def signup(request):
    if request.method == "POST":
        form = Register(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get("username")
            first_name = form.cleaned_data.get("first_name")
            last_name = form.cleaned_data.get("last_name")
            email = form.cleaned_data.get("email")
            password = form.cleaned_data.get("password1")
            form.save()
            messages.success(request, f'Account has been created for {username}. Continue to Log in.')
            new_user = authenticate(username=username, password=password, first_name=first_name, last_name=last_name, email=email)
            if new_user is not None:
                login(request, new_user)
                return redirect('user-login')


    else:
        form = Register()

    context = {
        'form' : form,
    }
    return render(request, 'signup.html', context)

def index(request):
    user = request.user
    form = TicketForm()
    tickets = Ticket.objects.all()
    tickets_by_user = Ticket.objects.filter(Q(user=user) | Q(assigned_to=user.username)).count()
    recent_tickets = Ticket.objects.filter(user=user).order_by('-created_at')[:10]
    overall_recent_tickets = Ticket.objects.all().order_by('-created_at')[:10]

    current_year = timezone.now().year
    tickets_by_month = Ticket.objects.filter(created_at__year=current_year).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(count=Count('ticket_number')).order_by('month').filter(Q(user=user) | Q(assigned_to=user.username))

    months = [datetime(2000, i + 1, 1).strftime('%b') for i in range(12)]
    ticket_counts = [0] * 12

    for order in tickets_by_month:
        month_index = order['month'].month - 1
        ticket_counts[month_index] = order['count']

    overall_tickets_by_month = Ticket.objects.filter(created_at__year=current_year).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(count=Count('ticket_number')).order_by('month').all()

    months = [datetime(2000, i + 1, 1).strftime('%b') for i in range(12)]
    overall_ticket_counts = [0] * 12

    for order in overall_tickets_by_month:
        month_index = order['month'].month - 1
        overall_ticket_counts[month_index] = order['count']

    overall_total_tickets = Ticket.objects.all().count()
    overall_resolved_tickets = Ticket.objects.all().count()
    overall_assigned_tickets = Ticket.objects.filter(status='Assigned').count()

    today = timezone.now().date()
    # assigned_to_today = Ticket.objects.filter( user=user, updated_at__date=today).count()
    user_tickets_today = Ticket.objects.filter(Q(user=user) | Q(assigned_to=user.username), created_at__date=today, updated_at__date=today ).count()
    overall_tickets_today = Ticket.objects.filter(created_at__date=today).count()
    assigned_to = Ticket.objects.filter( user=user, status='Assigned' or 'In Progress').count()
    total_resolved = Ticket.objects.filter(user=user, status='Done').count()

    context = {
        'overall_recent_tickets' : overall_recent_tickets,
        'overall_assigned_tickets' : overall_assigned_tickets,
        'overall_total_tickets' : overall_total_tickets,
        'overall_resolved_tickets' : overall_resolved_tickets,
        'overall_tickets_today' : overall_tickets_today,
        'user' : user,
        'form' : form,
        'tickets' : tickets,
        'tickets_by_user' : tickets_by_user,
        'user_tickets_today' : user_tickets_today,
        'assigned_to' : assigned_to,
        'total_resolved' : total_resolved,
        'months' : months,
        'ticket_counts' : ticket_counts,
        'overall_ticket_counts' : overall_ticket_counts,
        'recent_tickets' : recent_tickets

    }
    return render(request, 'index.html', context)

def ticketlogs(request):
    user = request.user
    tickets = Ticket.objects.annotate(
        month=ExtractMonth('created_at'),
        day=ExtractDay('created_at'),
        year=ExtractYear('created_at')
    ).filter(
        Q(user=user) | Q(assigned_to=user.username) | Q(created_by=user.username)
    ).order_by('-id')

    paginator = Paginator(tickets, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    overall_tickets = Ticket.objects.annotate(
        month=ExtractMonth('created_at'),
        day=ExtractDay('created_at'),
        year=ExtractYear('created_at')
    ).order_by('-id')

    paginator = Paginator(overall_tickets, 15)
    page_number2 = request.GET.get('page')
    page_obj2 = paginator.get_page(page_number2)

    all_users = User.objects.all()

    context = {
        'tickets': tickets,
        'all_users': all_users,
        'page_obj': page_obj,
        'page_obj2' : page_obj2
    }

    return render(request, 'ticketlogs.html', context)


def faqs(request):
    return render(request, 'faqs.html')


def search_ticket(request):
    if request.method == 'POST':
        form = TicketSearchForm(request.POST)
        if form.is_valid():
            id_number = form.cleaned_data.get('id_number')
            tickets = Ticket.objects.filter(id_number=id_number).order_by('-created_at')

            # Render the search results to an HTML string
            html = render_to_string('partials/search_results.html', {'tickets': tickets}, request=request)
            return JsonResponse({'html': html})
    return JsonResponse({'html': '<p>No tickets found. Search for ticket history using ID Number.</p>'})


def create_ticket(request):
    if request.method == 'POST':
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.user = request.user  # Associate the logged-in user with the `user` field
            ticket.save()
            messages.success(request, 'Ticket has been created successfully')
            return redirect('dashboard')  # Redirect to the dashboard or any other view
    else:
        form = TicketForm()

    return render(request, 'ticket_form_modal.html', {'form': form})




def update_ticket_status(request, pk):
    if request.method == 'POST':
        ticket = get_object_or_404(Ticket, id=pk)

        new_status = request.POST.get('status')
        assigned_user_username = request.POST.get('assigned_user')

        status_changed = new_status and new_status != ticket.status
        user_changed = (assigned_user_username != ticket.assigned_to) if assigned_user_username else ticket.assigned_to is not None

        if user_changed:
            if assigned_user_username:
                assigned_user = get_object_or_404(User, username=assigned_user_username)
                ticket.assigned_to = assigned_user_username
                ticket.status = 'Assigned'

                # Create notification for assigned user
                Notification.objects.create(
                    user=assigned_user,
                    ticket=ticket,
                    message=f'Ticket #{ticket.ticket_number} has been assigned to you'
                )

                messages.success(request, f'Ticket #{ticket.ticket_number} has been assigned to {assigned_user_username}')
            else:
                ticket.assigned_to = None
                messages.success(request, f'Ticket #{ticket.ticket_number} has been unassigned')

        if status_changed:
            previous_status = ticket.status
            ticket.status = new_status

            # Create notification for ticket creator
            if new_status in ['Done', 'In Progress']:
                Notification.objects.create(
                    user=ticket.user,
                    ticket=ticket,
                    message=f'Your ticket #{ticket.ticket_number} status has been changed to {new_status}'
                )

            messages.success(request, f'Ticket #{ticket.ticket_number} status changed from {previous_status} to {new_status}')

        ticket.save()
        return redirect('ticketlogs')

    return HttpResponseRedirect('/')





def searchdata(request):
    user = request.user
    q = request.GET.get('query')

    if q:
        if request.user.is_superuser:
            # Superusers can search all tickets
            tickets = Ticket.objects.filter(
                Q(ticket_number__icontains=q) |
                Q(id_number__icontains=q) |
                Q(assigned_to__icontains=q) |
                Q(user__username__icontains=q) |  # Search by creator's username
                Q(status__icontains=q) |
                Q(category__icontains=q) |
                Q(client_type__icontains=q) |
                Q(mode__icontains=q)
            ).order_by('-id')
        else:
            # Regular users can only search their own tickets and tickets assigned to them
            tickets = Ticket.objects.filter(
                Q(user=user) | Q(assigned_to=user.username)
            ).filter(
                Q(ticket_number__icontains=q) |
                Q(id_number__icontains=q) |
                Q(assigned_to__icontains=q) |
                Q(status__icontains=q) |
                Q(category__icontains=q) |
                Q(client_type__icontains=q) |
                Q(mode__icontains=q)
            ).order_by('-id')
    else:
        # If no search query, show all tickets for superuser or user's tickets for regular users
        if request.user.is_superuser:
            tickets = Ticket.objects.all().order_by('-id')
        else:
            tickets = Ticket.objects.filter(
                Q(user=user) | Q(assigned_to=user.username)
            ).order_by('-id')

    # Paginate the results
    paginator = Paginator(tickets, 15)  # Show 15 tickets per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # For superusers, we'll use page_obj2 to maintain consistency with the template
    if request.user.is_superuser:
        page_obj2 = page_obj
        page_obj = None
    else:
        page_obj2 = None

    context = {
        'page_obj': page_obj,
        'page_obj2': page_obj2,
        'query': q,
        'all_users': User.objects.all(),  # Add this for the assign dropdown
    }
    return render(request, 'ticketlogs.html', context)

def is_valid_queryparam(param):
    return param != '' and param is not None

def report(request):
    user = request.user

    if request.user.is_superuser:
        tickets = Ticket.objects.annotate(
            month_year=TruncMonth('created_at')
        ).values('month_year').annotate(
            count=Count('id')
        ).order_by('-month_year')
    else:
        tickets = Ticket.objects.filter(Q(user=user) | Q(assigned_to=user.username)).annotate(
            month_year=TruncMonth('created_at')
        ).values('month_year').annotate(
            count=Count('id')
        ).order_by('-month_year')

    status_options = Ticket.objects.values_list('status', flat=True).distinct()
    # pending_orders = Ticket.objects.filter( status='pending').count()
    available_months = [(ticket['month_year'].strftime('%B %Y'), ticket['month_year'].strftime('%Y-%m-01'), ticket['month_year'].strftime('%Y-%m-31')) for ticket in tickets]
    mode_options = Ticket.objects.values_list('mode', flat=True).distinct()
    tl = Ticket.objects.order_by('-id')
    tickets1 = Ticket.objects.all()
    tickets2 = Ticket.objects.filter(Q(user=user) | Q(assigned_to=user.username)).order_by('-id')

    user_set = set()
    assigned_to_set = set()

    for ticket in tickets1:
        if ticket.user:
            user_set.add(ticket.user)
        if ticket.assigned_to:
            assigned_to_set.add(ticket.assigned_to)

    ticket_id = request.GET.get('id')
    ticket_number = request.GET.get('ticket_number')
    created_by = request.GET.get('created_by')
    id_number = request.GET.get('id_number')
    status = request.GET.get('status')
    mode = request.GET.get('mode')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    assigned_to = request.GET.get('assigned_to')

    request.session['id'] = ticket_id
    request.session['ticket_number'] = ticket_number
    request.session['created_by'] = created_by
    request.session['id_number'] = id_number
    request.session['date_from'] = date_from
    request.session['date_to'] = date_to
    request.session['status'] = status
    request.session['mode'] = mode
    request.session['assigned_to'] = assigned_to

    if request.user.is_superuser:
        tl = tl.filter(
            Q(created_at__range=[date_from, date_to]) if date_from and date_to else Q()
        )
        if is_valid_queryparam(ticket_id):
            tl = tl.filter(id=ticket_id)
        if is_valid_queryparam(ticket_number):
            tl = tl.filter(ticket_number__icontains=ticket_number)
        if is_valid_queryparam(created_by):
            try:
                user = User.objects.get(username=created_by)  # Retrieve the user by username
                tl = tl.filter(Q(user=user))  # Filter by user object
            except User.DoesNotExist:
                tl = tl.none()
        if is_valid_queryparam(id_number):
            tl = tl.filter(id_number__icontains=id_number)
        if is_valid_queryparam(status):
            tl = tl.filter(status=status)
        if is_valid_queryparam(mode):
            tl = tl.filter(mode=mode)
        if is_valid_queryparam(assigned_to):
            tl = tl.filter(assigned_to__icontains=assigned_to)

        page = request.GET.get('page', 1)
        paginator = Paginator(tl, 30)

        try:
            tl = paginator.page(page)
        except PageNotAnInteger:
            tl = paginator.page(1)
        except EmptyPage:
            tl = paginator.page(paginator.num_pages)
    else:
        tickets2 = tickets2.filter(
        Q(created_at__range=[date_from, date_to]) if date_from and date_to else Q()
        )
        if is_valid_queryparam(ticket_id):
            tickets2 = tickets2.filter(id=ticket_id)
        if is_valid_queryparam(ticket_number):
            tickets2 = tickets2.filter(ticket_number__icontains=ticket_number)
        if is_valid_queryparam(created_by):
            try:
                user = User.objects.get(username=created_by)  # Retrieve the user by username
                tickets2 = tickets2.filter(Q(user=user))  # Filter by user object
            except User.DoesNotExist:
                tickets2 = tickets2.none()
        if is_valid_queryparam(id_number):
            tickets2 = tickets2.filter(id_number__icontains=id_number)
        if is_valid_queryparam(status):
            tickets2 = tickets2.filter(status=status)
        if is_valid_queryparam(mode):
            tickets2 = tickets2.filter(mode=mode)
        if is_valid_queryparam(assigned_to):
            tickets2 = tickets2.filter(assigned_to__icontains=assigned_to)
        page = request.GET.get('page', 1)
        paginator = Paginator(tickets2, 30)

        try:
            tickets2 = paginator.page(page)
        except PageNotAnInteger:
            tickets2 = paginator.page(1)
        except EmptyPage:
            tickets2 = paginator.page(paginator.num_pages)



    context = {
        'tickets1' : tickets,
        'tickets2' : tickets2,
        'ticket_list': tl,
        'ticket_id' : ticket_id,
        'created_by': created_by,
        'assigned_to' : assigned_to,
        'ticket_number': ticket_number,
        'mode': mode,
        'date_from': date_from,
        'date_to': date_to,
        'status': status,
        'available_months': available_months,
        'user_options': user_set,
        'assigned_to_options': assigned_to_set,
        'status_options' : status_options,
        'mode_options' : mode_options
    }
    return render(request, 'report.html', context)

@login_required
def ticket_excel(request):
    tl = Ticket.objects.order_by('user')

    ticket_number = request.session.get('ticket_number')
    created_by = request.session.get('created_by')
    mode = request.session.get('mode')
    id_number = request.session.get('id_number')
    date_from = request.session.get('date_from')
    date_to = request.session.get('date_to')
    status = request.session.get('status')
    assigned_to = request.session.get('assigned_to')

    # Apply filters to the queryset
    if date_from and date_to:
        tl = tl.filter(created_at__range=[date_from, date_to])

    if is_valid_queryparam(ticket_number):
        tl = tl.filter(ticket_number__icontains=ticket_number)
    if is_valid_queryparam(created_by):
        tl = tl.filter(user__username=created_by)
    if is_valid_queryparam(id_number):
        tl = tl.filter(id_number__icontains=id_number)
    if is_valid_queryparam(status):
        tl = tl.filter(status=status)
    if is_valid_queryparam(mode):
        tl = tl.filter(mode=mode)
    if is_valid_queryparam(assigned_to):
        tl = tl.filter(assigned_to=assigned_to)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Ticket Report {date_from} to {date_to}.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.merge_cells('A1:I1')
    worksheet.merge_cells('A2:I2')
    first_cell = worksheet.cell(row=1, column=1)
    first_cell.value = f"Report For Tickets Generated on {timezone.now()}"

    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = f'Request List {date_from} to {date_to}'

    columns = ['Ticket ID', 'Created By','Assigned To', 'ID Number', 'Client Type','Category', 'Status', 'Date Created', 'Updated At' ]
    row_num = 3

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = max(len(str(column_title)), 12)

    for ticket in tl:
        row_num += 1
        row = [
            ticket.ticket_number,
            str(ticket.user.username),  # Convert User object to username string
            ticket.assigned_to,
            ticket.id_number,
            ticket.client_type,
            ticket.category,
            ticket.status,
            ticket.created_at.replace(tzinfo=None) if ticket.created_at else None,
            ticket.updated_at.replace(tzinfo=None) if ticket.updated_at else None,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if isinstance(cell_value, datetime):
                cell.number_format = 'yyyy-mm-dd HH:MM:SS'
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(worksheet.column_dimensions[column_letter].width, len(str(cell_value)) + 2)

        worksheet.row_dimensions[row_num].height = 14.4

    total_count_cell = worksheet.cell(row=row_num + 1, column=1)
    total_count_cell.value = "Total Count"
    total_count_cell.font = Font(bold=True)
    total_count_cell.alignment = Alignment(horizontal="center", vertical="center")

    total_count_value_cell = worksheet.cell(row=row_num + 1, column=2)
    total_count_value_cell.value = len(tl)
    total_count_value_cell.font = Font(bold=True)
    total_count_value_cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(response)
    return response


# @login_required
# def order_pdf(request):
#     ol = Order.objects.order_by('id')

#     # Retrieve filters from session
#     order_id = request.session.get('id')
#     name = request.session.get('name')
#     issued_to = request.session.get('issued_to')
#     itemName = request.session.get('item_name')
#     date_from = request.session.get('date_from')
#     date_to = request.session.get('date_to')
#     status = request.session.get('status')
#     released_by = request.session.get('released_by')
#     received_by = request.session.get('received_by')

#     # Apply filters to queryset
#     ol = ol.filter(
#         Q(date__range=[date_from, date_to]) if date_from and date_to else Q()
#     )
#     if is_valid_queryparam(order_id):
#         ol = ol.filter(id=order_id)
#     if is_valid_queryparam(name):
#         ol = ol.filter(users__username__icontains=name)
#     if is_valid_queryparam(issued_to):
#         ol = ol.filter(issued_to__icontains=issued_to)
#     if is_valid_queryparam(itemName):
#         ol = ol.filter(item_name__name__icontains=itemName)
#     if is_valid_queryparam(status):
#         ol = ol.filter(status=status)
#     if is_valid_queryparam(released_by):
#         ol = ol.filter(released_by__icontains=released_by)
#     if is_valid_queryparam(received_by):
#         ol = ol.filter(returned_to__icontains=received_by)

#     # Create a PDF report
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="Inventory_Request_Report.pdf"'

#     data = []
#     header = ['Request ID', 'Issued By', 'Issued To', 'Item Name', 'Quantity', 'Date Created', 'Date Received', 'Status', 'Released By', 'Received By']
#     data.append(header)
#     for order in ol:
#         data.append([order.id, order.users.username, order.issued_to, order.item_name.name, order.request_quantity,
#                      order.date.replace(tzinfo=None) if order.date else None,
#                      order.returned_date.replace(tzinfo=None) if order.returned_date else None,
#                      order.status, order.released_by, order.returned_to])

#     total_count = len(ol)

#     doc = SimpleDocTemplate(response, pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)

#     table_style = TableStyle([
#         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#         ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
#         ('FONTSIZE', (0, 0), (-1, -1), 10),
#         ('GRID', (0, 0), (-1, -1), 1, colors.black),
#     ])

#     table = Table(data, style=table_style)
#     styles = getSampleStyleSheet()

#     elements = []
#     elements.append(Paragraph(f"A Report of Requests Generated on {timezone.now()}", styles['title']))
#     elements.append(Paragraph(f"Filters used:", styles['title']))
#     elements.append(Paragraph(f"Request ID: {order_id if order_id else 'All'}, Name: {name if name else 'All'}, Item Name: {itemName if itemName else 'All'}, Date From: {date_from if date_from else 'All'}, Date To: {date_to if date_to else 'All'}, Status: {status if status else 'All'}, Released By: {released_by if released_by else 'All'}, Received By: {received_by if received_by else 'All'}", styles['Normal']))
#     elements.append(table)
#     elements.append(Paragraph(f"Total Count of Requests: {total_count}", styles['Normal']))

#     doc.build(elements)
#     return response

@login_required
def edit_ticket(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)

    # Check if the user has permission to edit the ticket
    if not request.user.is_superuser and request.user != ticket.user:
        messages.error(request, "You don't have permission to edit this ticket.")
        return redirect('ticketlogs')

    if request.method == 'POST':
        # Get form data
        client_type = request.POST.get('client_type')
        id_number = request.POST.get('id_number')
        category = request.POST.get('category')
        description = request.POST.get('description')

        # Validate the data
        if not all([client_type, id_number, category, description]):
            messages.error(request, 'All fields are required.')
            return redirect('ticketlogs')

        # Validate choices
        if client_type not in [choice[0] for choice in Ticket.CLIENT_TYPE]:
            messages.error(request, 'Invalid client type.')
            return redirect('ticketlogs')

        if category not in [choice[0] for choice in Ticket.CATEGORY]:
            messages.error(request, 'Invalid category.')
            return redirect('ticketlogs')

        # Update ticket
        ticket.client_type = client_type
        ticket.id_number = id_number
        ticket.category = category
        ticket.description = description
        ticket.save()

        messages.success(request, f'Ticket #{ticket.ticket_number} has been updated successfully')
        return redirect('ticketlogs')

    return redirect('ticketlogs')

@login_required
def delete_ticket(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)

    # Check if the user has permission to delete the ticket
    if not request.user.is_superuser and request.user != ticket.user:
        messages.error(request, "You don't have permission to delete this ticket.")
        return redirect('ticketlogs')

    if request.method == 'POST':
        ticket_number = ticket.ticket_number
        ticket.delete()
        messages.success(request, f'Ticket #{ticket_number} has been deleted successfully')

    return redirect('ticketlogs')

def get_notifications(request):
    if request.user.is_authenticated:
        notifications = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).order_by('-created_at')[:5]  # Get 5 most recent unread notifications

        context = {
            'notifications': notifications,
            'unread_count': notifications.count()
        }

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render_to_string('partials/notifications.html', context, request=request)
            return JsonResponse({'html': html, 'count': notifications.count()})

        return render(request, 'partials/notifications.html', context)
    return JsonResponse({'html': '', 'count': 0})

def mark_notification_read(request, notification_id):
    if request.method == 'POST' and request.user.is_authenticated:
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})
